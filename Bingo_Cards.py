import time
import random
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side


def genrandnum():
    global ele
    ele = random.choice(num)


def gencard():
    global num, bc, ele, testele
    for i in range(1, 26):
        num.append(i)
    for i in range(5):
        bc.append([])
        while len(bc[i]) <= 4:
            genrandnum()
            if ele not in testele:
                testele.append(ele)
                bc[i].append(ele)
            else:
                genrandnum()
    genbcdict()


def genbcdict():
    global ch, df, bc, bcdict
    hl = list("BINGO")
    for i in range(5):
        bcdict[hl[i]] = bc[i]
    df = pd.DataFrame(bcdict)
    if ch == 1:
        newbc()
        ch = 2
    else:
        savebc()


def savebc():
    global df
    try:
        writer = pd.ExcelWriter('bingo_cards.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.workbook = load_workbook('bingo_cards.xlsx')
        writer.worksheets = dict((ws.title, ws) for ws in writer.workbook.worksheets)
        reader = pd.read_excel('bingo_cards.xlsx')
        if len(reader) == 0:
            df.to_excel(writer, index=False)
        else:
            df.to_excel(writer, index=False, startrow=len(reader) + 1)
        writer.close()
        print(df)
        formatdata()
        resetdata()
    except FileNotFoundError:
        newbc()


def newbc():
    df.to_excel('bingo_cards.xlsx', index=False)
    print(df)
    formatdata()
    resetdata()


def formatdata():
    reader = pd.read_excel('bingo_cards.xlsx')
    wb = load_workbook('bingo_cards.xlsx')
    ws = wb.active
    cell = ws['A1':'E' + str(len(reader) + 1)]
    for r in cell:
        for c in r:
            c.font = Font(size=18)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for i in range(1, len(reader), 6):
        hc = ws['A' + str(i):'E' + str(i)]
        for r in hc:
            for c in r:
                c.font = Font(size=22, bold=True)
    for r in range(1, len(reader) + 2):
        ws.row_dimensions[r].height = 50
    cl = list("ABCDE")
    for c in cl:
        ws.column_dimensions[c].width = 16
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 11
    ws.page_margins.top = 0.7519685
    ws.page_margins.bottom = 0.7519685
    ws.page_margins.left = 0.7007874
    ws.page_margins.right = 0.7007874
    ws.page_margins.header = 0.299213
    ws.page_margins.footer = 0.299213
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.oddHeader.left.text = "ID: &[Page]"
    ws.oddFooter.center.text = "CREATED BY: VIVEK"
    ws.evenHeader.left.text = "ID: &[Page]"
    ws.evenFooter.center.text = "CREATED BY: VIVEK"
    wb.save('bingo_cards.xlsx')


def resetdata():
    global ele, noc, num, bc, testele, bcdict, df
    ele = 0
    noc = 0
    num = list()
    bc = list()
    testele = list()
    bcdict = dict()
    df = pd.DataFrame()


def getnoc():
    global ch, noc
    try:
        ch = int(input("\nOptions:\n1. Generate New Cards\n2. Keep Existing Cards\n0. Exit\nChoice (Press 1, 2 or 0): "))
        if ch in [1,2]:
            noc = int(input("How many cards do you want to generate: "))
            for i in range(noc):
                print("\nGenerating bingo card", i + 1)
                time.sleep(0.25)
                gencard()
                time.sleep(0.25)
                print("Bingo card", i + 1, "saved sucessfully...")
            getnoc()
        elif ch == 0:
            exit()
        else:
            print("Invalid input...")
            getnoc()
    except ValueError:
        print("Invalid input...")
        getnoc()


ch = 0
ele = 0
noc = 0
num = list()
bc = list()
testele = list()
bcdict = dict()
df = pd.DataFrame()
getnoc()
